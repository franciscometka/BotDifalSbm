"""
Geracao da planilha de controle de DIFAL usando openpyxl.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

COLUNAS = [
    "Status",
    "Número NF",
    "Série",
    "Chave NF-e",
    "Data emissão",
    "Cliente",
    "CPF/CNPJ cliente",
    "UF destino",
    "Valor total NF",
    "Valor ICMS UF destino",
    "Valor FCP UF destino",
    "Valor ICMS UF remetente",
    "CFOPs",
    "NCMs",
    "Precisa DIFAL?",
    "Portal sugerido",
    "PDF Guia",
    "Status da Guia",
    "Observações",
]


def gerar_planilha(linhas, caminho_saida="controle_difal.xlsx"):
    """
    Gera o arquivo Excel de controle a partir de uma lista de dicts.

    Cada item de `linhas` deve ter as chaves usadas em `COLUNAS`
    (em snake_case, ver mapeamento abaixo).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle DIFAL"

    ws.append(COLUNAS)
    cabecalho_fonte = Font(bold=True, color="FFFFFF")
    cabecalho_preenchimento = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for coluna in range(1, len(COLUNAS) + 1):
        celula = ws.cell(row=1, column=coluna)
        celula.font = cabecalho_fonte
        celula.fill = cabecalho_preenchimento

    for linha in linhas:
        ws.append([
            linha.get("status", ""),
            linha.get("numero_nf", ""),
            linha.get("serie", ""),
            linha.get("chave_acesso", ""),
            linha.get("data_emissao", ""),
            linha.get("nome_destinatario", ""),
            linha.get("doc_destinatario", ""),
            linha.get("uf_destino", ""),
            linha.get("valor_total_nf", ""),
            linha.get("valor_icms_uf_dest", ""),
            linha.get("valor_fcp_uf_dest", ""),
            linha.get("valor_icms_uf_remet", ""),
            linha.get("cfops", ""),
            linha.get("ncms", ""),
            linha.get("precisa_difal", ""),
            linha.get("portal_sugerido", ""),
            linha.get("pdf_guia", ""),
            linha.get("status_guia", ""),
            linha.get("observacoes", ""),
        ])

    larguras = [10, 12, 8, 46, 20, 30, 18, 10, 14, 16, 16, 18, 16, 16, 14, 22, 22, 18, 30]
    for indice, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = largura

    ws.freeze_panes = "A2"

    wb.save(caminho_saida)
    return caminho_saida
